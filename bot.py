import asyncio
import logging
import os
import uuid
import json
from datetime import datetime

from aiogram import Bot, Dispatcher
from aiogram.types import (
    Message, ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    CallbackQuery, FSInputFile
)
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import InputMediaPhoto
from aiogram.utils.keyboard import InlineKeyboardBuilder
from dotenv import load_dotenv
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from openpyxl import Workbook, load_workbook
from calendar_utils import build_calendar, month_title
from datetime import date, datetime, timedelta
from admin.handlers import router as admin_router
from db import (
    init_db,
    save_order,
    sign_contract,
    get_last_booking_by_user,
    mark_paid,
    init_calendar_for_month,
    get_available_dates_range,
    book_places
)
# ======================
# НАСТРОЙКИ
# ======================

load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
logging.basicConfig(level=logging.INFO)

bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())
dp.include_router(admin_router)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "orders.xlsx")

# ======================
# 🔥 НАСТРОЙКИ КАЛЕНДАРЯ
# ======================

MAX_DAYS_AHEAD = 14

BLOCKED_DATES = {
    # пример:
    # "2026-01-10",
    # "2026-01-18"
}

# ======================
# ТЕКСТЫ (FAQ / INFO)
# ======================
FAQ_TEXT = (
    " <b>⁉️Часто задаваемые вопросы</b>\n\n"
    "<b>1) Как проходит процедура выбора и бронирования экскурсии?</b>\n\n"
    "☑️️Вы выбираете экскурсию в боте\n"
    "☑️️Указываете дату, ФИО, телефон и количество человек для бронирования\n"
    "☑️️Мы фиксируем бронирование, рассчитываем стоимость\n"
    "☑️️Вы знакомитесь и подписываете договор по оказанию услуг прямо в боте\n"
    "☑️️В день экскурсии приезжаете к месту подачи и оплачиваете поездку\n\n"
    "<b>2) В каком формате проходит экскурсия?</b>\n\n"
    "Вас встречает водитель по адресу места подачи, указанному в договоре. Поездка по маршруту осуществляется на комфортабельном транспорте в сопровождении аудиогида. По прибытии в пункт назначения у Вас будет время посетить достопримечательность, осмотреться и сделать фотографии в течение заранее оговоренного времени. Водитель будет ожидать Вас в машине.\n\n"
    "<b>3) Нужно ли вносить предоплату?</b>\n\n"
    "Нет, вы можете оплатить поездку на месте — наличными или переводом.\n\n"
    "<b>4) Где находится точка сбора и подачи машины в день экскурсии?</b>\n\n"
    "Место подачи машины будет указано в договоре фрахтования, но, как правило, мы забираем пассажиров со станции МЦД Опалиха со стороны ЖК Опалиха О3. Адрес: Московская обл., Красногорск, станция Опалиха. https://go.2gis.com/2Q7no \n\n"
    "<b>5) Какой транспорт используется?</b>\n\n"
    "Комфортабельный минивэн GAC M8, категория M1.\n\n"
    "<b>6) Можно ли отменить поездку?</b>\n\n"
    "Да, просто свяжитесь с нами заранее по номеру 88002017938.\n\n"
    "<b>7) Как связаться с водителем в день экскурсии</b>\n\n"
    "Позвоните по номеру 88002017938, оператор укажет контактные данные водителя\n\n"
    "<b>8) Задать любой вопрос по работе нашего сервиса</b>\n\n"
    "Пожалуйста, свяжитесь с нами по номеру 88002017938\n\n"

)

ABOUT_TEXT = (
    "🚐 <b>Комфортные экскурсии и поездки</b>\n"
    "Мы организуем автомобильные экскурсии на современных минивэнах по Москве и Московской области.\n\n"
    "🛡 <b>Безопасность</b> — продуманные маршруты и надёжный транспорт\n"
    "💺 <b>Комфорт</b> — просторный салон, удобные сиденья и приятная атмосфера\n"
    "✨ <b>Впечатления</b> — поездки, которые хочется вспоминать\n"
    "👨‍💻 <b>Сервис</b> - профессиональные и вежливые водители, автоматизированная система быстрого бронирования, поддержка клиентов на всех этапах\n\n"
    "🛣️ <b>GACVoyage</b> - комфорт и яркие эмоции с каждым километром 💙"
)

REVIEWS = [
    {
        "photo": "images/review_1.jpg",
        "text": (
            "⭐️⭐️⭐️⭐️⭐️\n"
            "<b>Анна, Москва</b>\n\n"
            "«Ездили всей семьёй — двое детей, коляска, рюкзаки.\n"
            "Машина чистая и просторная, водитель помог с посадкой.\n"
            "Поездка прошла спокойно и без суеты» 💬"
        )
    },
    {
        "photo": "images/review_2.png",
        "text": (
            "⭐️⭐️⭐️⭐️⭐️\n"
            "<b>Дмитрий, Санкт-Петербург</b>\n\n"
            "«Водитель приехал заранее, аккуратно вёл машину\n"
            "и рассказал интересные факты по дороге.\n"
            "Это была не просто поездка, а впечатление»"
        )
    },
    {
        "photo": "images/review_3.png",
        "text": (
            "⭐️⭐️⭐️⭐️⭐️\n"
            "<b>Екатерина, Казань</b>\n\n"
            "«После поездки осталось ощущение заботы.\n"
            "Всё продумано до мелочей. Очень рекомендую!» 💙"
        )
    }
]

CONTACT_TEXT = (
    "📞 <b>Связаться с нами</b>\n\n"
    "Федеральный номер:\n"
    "<b>8-800-201-79-38</b>\n\n"
    "Мы на связи ежедневно c 10:00 до 22:00."
)
# ======================
# ВОДИТЕЛИ
# ======================

DRIVERS = [
    {
        "name": "Диспетчер Слава",
        "telegram_id": 292972793  # ← ЗАМЕНИ на реальный Telegram ID
    },
    {
        "name": "Сергей",
        "telegram_id": 335639358 #292972793  # ← ЗАМЕНИ на реальный Telegram ID
    }
]

# ======================
# ДАННЫЕ ЭКСКУРСИЙ
# ======================

EXCURSIONS = [
    {
        "id": "pilgrims",
        "title": "Паломники (Храмы и церкви России)",
        "description": (
            "Посещение трёх храмов:\n"
            "✅ Ново-Иерусалимский (Истра)\n"
            "✅ Храм вооружённых сил России (Кубинка)\n"
            "✅ Храм Христа Спасителя (Москва)\n\n"
            "⏰ В каждом храме свободное время 40-60 минут\n"
            "⏰ Выезд в 08:00\n"
            "⏰ Длительность ~4-5 часов\n\n"
            "💵 Стоимость: 4000 ₽/чел."
        ),
        "images": [
            "media/piligrims/5389031389581217457.jpg",
            "media/piligrims/5389031389581217458.jpg",
            "media/piligrims/5389031389581217459.jpg",
            "media/piligrims/5389031389581217460.jpg",
            "media/piligrims/5389031389581217464.jpg",
            "media/piligrims/5389031389581217470.jpg",
            "media/piligrims/5389031389581217472.jpg",
            "media/piligrims/5389031389581217473.jpg",
            "media/piligrims/5389031389581217474.jpg",
            "media/piligrims/5389031389581217475.jpg",
            "media/piligrims/5389031389581217476.jpg",
            "media/piligrims/Изображение PNG 10.png",
        ],
        "start_time": "08:00",
        "price": 4000,
        "prepayment_percent": 30,
        "pickup": {
            "title": "ст. Опалиха (со стороны ЖК Опалиха О3)",
            "address": "Московская обл., Красногорск, станция Опалиха",
            "gis": "https://go.2gis.com/2Q7no"
        },
        "route": [
            {
                "name": "Ново-Иерусалимский монастырь",
                "address": "Московская обл., г. Истра, ул. Советская, 2"
            },
            {
                "name": "Храм Вооружённых сил РФ",
                "address": "Московская обл., Одинцовский г.о., парк Патриот"
            },
            {
                "name": "Храм Христа Спасителя",
                "address": "г. Москва, ул. Волхонка, 15"
            }
        ]
    },
    {
        "id": "new_year",
        "title": "Новый год, 2026",
        "description": (
            "🏮Посещение самых красивых новогодних локаций Москвы для фотосессий 📸:\n"
            "✅ Парад новогодних ёлок у ЦУМа 🎄\n"
            "✅ Большой театр 🏛️\n"
            "✅ Дом графа Шереметьева 🏰\n"
            "✅ Парк Акведук 🎡\n\n"
            "⏰ Выезд в 15:00\n\n "
            "💵 Cтоимость: 3500 ₽/чел"
        ),
        "images": [
            "media/new_year/5389031389581217210.jpg",
            "media/new_year/5389031389581217212.jpg",
            "media/new_year/5389031389581217213.jpg",
            "media/new_year/5389031389581217214.jpg",
            "media/new_year/5389031389581217215.jpg",
            "media/new_year/5389031389581217216.jpg",
            "media/new_year/5389031389581217217.jpg",
            "media/new_year/5389031389581217218.jpg",
            "media/new_year/5389031389581217226.jpg",
            "media/new_year/5389031389581217229.jpg",
        ],
        "start_time": "15:00",
        "price": 3500,
        "prepayment_percent": 30,
        "pickup": {
            "title": "ст. Опалиха (со стороны ЖК Опалиха О3)",
            "address": "Московская обл., Красногорск, станция Опалиха",
            "gis": "https://go.2gis.com/2Q7no"
        },
        "route": [
            {
                "name": "Парад новогодних ёлок у ЦУМа",
                "address": "г. Москва, ул. Петровка, 2"
            },
            {
                "name": "Большой театр",
                "address": "г. Москва, Театральная площадь, 1"
            },
            {
                "name": "Дом графа Шереметьева",
                "address": "г. Москва, Романов переулок, 2"
            },
            {
                "name": "Парк Акведук",
                "address": "г. Москва, ул. Малахитовая"
            }
        ]
    }
]


EXECUTOR = {
    "name": "ИП Шин Сергей Тимофеевич",
    "inn": "621200217989",
    "city": "Москва",
    "car": "GAC M8",
    "car_type": "M1",
    "plate": "В412УМ62"
}

# ======================
# FSM
# ======================

class BookingStates(StatesGroup):
    choose_excursion = State()
    date = State()
    name = State()
    phone = State()
    pickup = State()   # ★ НОВОЕ
    count = State()
    payment = State()

# ======================
# КЛАВИАТУРЫ
# ======================

main_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🚗 Выбрать экскурсию")],
        [KeyboardButton(text="⭐ Отзывы"), KeyboardButton(text="ℹ️ О нас")],
        [KeyboardButton(text="❓ FAQ"), KeyboardButton(text="📞 Связаться")]
    ],
    resize_keyboard=True
)

def excursion_kb():
    buttons = [[KeyboardButton(text=ex["title"])] for ex in EXCURSIONS]
    buttons.append([KeyboardButton(text="⬅️ Назад")])
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

def book_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 Забронировать", callback_data="start_booking")]
    ])

def payment_kb(amount: int):
    return InlineKeyboardMarkup(inline_keyboard=[
        #[InlineKeyboardButton(text=f"💳 Оплатить 30% ({amount} ₽)", callback_data="paid")],
        [InlineKeyboardButton(text="💰 Оплата на месте (картой или наличными)", callback_data="pay_on_place")],
        #[InlineKeyboardButton(text="✅ Я оплатил", callback_data="paid")]
    ])

view_contract_kb = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="📄 Ознакомиться с договором", callback_data="view_contract")]
])

sign_contract_kb = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="✍️ Подписать договор", callback_data="sign_contract")]
])

def calendar_kb(dates):
    kb = InlineKeyboardBuilder()

    for date, free in dates:
        day = date[-2:]
        kb.button(
            text=f"{day} ({free})",
            callback_data=f"date:{date}"
        )

    kb.adjust(3)
    return kb.as_markup()
# ======================
# PDF
# ======================

PAGE_WIDTH, PAGE_HEIGHT = A4

pdfmetrics.registerFont(TTFont("DejaVu", "DejaVuSans.ttf"))
pdfmetrics.registerFont(TTFont("DejaVu-Bold", "DejaVuSans-Bold.ttf"))

def generate_contract_pdf(order: dict, signed=False) -> str:
    filename = f"contract_{order['booking_id']}.pdf"

    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="TitleCenter",
        fontName="DejaVu-Bold",
        fontSize=13,
        alignment=TA_CENTER,
        spaceAfter=12
    ))

    styles.add(ParagraphStyle(
        name="Justify",
        fontName="DejaVu",
        fontSize=11,
        alignment=TA_JUSTIFY,
        leading=15,
        spaceAfter=10
    ))

    styles.add(ParagraphStyle(
        name="Section",
        fontName="DejaVu-Bold",
        fontSize=11,
        spaceBefore=15,
        spaceAfter=8
    ))

    story = []

    # ===== ЗАГОЛОВОК =====
    story.append(Paragraph(
        f"Договор фрахтования № {order['booking_id']}",
        styles["TitleCenter"]
    ))

    story.append(Paragraph(
        "транспортного средства для перевозки пассажиров по заказу",
        styles["TitleCenter"]
    ))

    # ===== ВВОДНЫЙ ТЕКСТ =====
    story.append(Paragraph(
        f"{EXECUTOR['name']}, ИНН {EXECUTOR['inn']}, именуемый в дальнейшем "
        f"Фрахтовщик, и {order['name']}, именуемый в дальнейшем Фрахтователь, "
        f"заключили настоящий договор о нижеследующем:",
        styles["Justify"]
    ))

    # ===== МАРШРУТ (ТОЛЬКО АДРЕСА) =====
    PICKUP_ADDRESSES = {
        "opalikha": "Московская обл., Красногорск, станция Опалиха",
        "tsum": "г. Москва, ул. Петровка, 2"
    }
    route_points = json.loads(order["route"])
    order["pickup_address"] = PICKUP_ADDRESSES["opalikha"]
    route_addresses = [
                          order["pickup_address"]
                      ] + [p["address"] for p in route_points]
    full_route = " — ".join(route_addresses)

    story.append(Paragraph(
        f"1.1. Фрахтовщик обязуется за плату в размере {order['price']} рублей "
        f"предоставить Фрахтователю всю вместимость транспортного средства для перевозки пассажиров и багажа.",
        styles["Justify"]
    ))

    story.append(Paragraph(
        f"<b>1.2. Место подачи:</b> {order['pickup_address']} (со стороны ЖК Опалиха О3)",
        styles["Justify"]
    ))

    story.append(Paragraph(
        f"<b>1.3. Маршрут перевозки:</b> {full_route}.",
        styles["Justify"]
    ))

    story.append(Paragraph(
        f"<b>1.4. Срок выполнения перевозки:</b> {order['date']}.",
        styles["Justify"]
    ))

    # ===== ТРАНСПОРТ =====
    story.append(Paragraph("1.5. Транспортное средство:", styles["Section"]))

    story.append(Paragraph(
        f"Марка и модель: {EXECUTOR['car']}<br/>"
        f"Тип ТС: {EXECUTOR['car_type']}<br/>"
        f"Государственный номер: {EXECUTOR['plate']}",
        styles["Justify"]
    ))

    # ===== РЕКВИЗИТЫ =====
    story.append(Spacer(1, 20))
    story.append(Paragraph("Реквизиты сторон", styles["TitleCenter"]))

    story.append(Paragraph(
        f"<b>Фрахтовщик:</b><br/>"
        f"{EXECUTOR['name']}<br/>"
        f"ИНН: {EXECUTOR['inn']}<br/>",
        # + ( "Подписано простой электронной подписью" if signed else "" ),
        styles["Justify"]
    ))

    story.append(Paragraph(
        f"<b>Фрахтователь:</b><br/>"
        f"ФИО: {order['name']}<br/>"
        f"Телефон: {order['phone']}<br/>",
       # + ( "Подписано простой электронной подписью" if signed else "" ),
        styles["Justify"]
    ))

    if signed:
        story.append(Spacer(1, 15))
        story.append(Paragraph(
            f"Подписано простой электронной подписью {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["Justify"]
        ))

    doc.build(story)
    return filename




    # ===== ОТМЕТКА О ПОДПИСИ =====
    if signed:
        y = base_y - 90
        c.setFont("DejaVu-Bold", 10)
        c.drawCentredString(
            PAGE_WIDTH / 2,
            y,
            f"Подписано ПЭП {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )

    c.save()
    return filename

# ======================
# Небольшой helper для получения экскурсии по id
# ======================
def get_excursion_by_id(excursion_id: str):
    return next((ex for ex in EXCURSIONS if ex["id"] == excursion_id), None)

# ======================
# Helper для отправки уведомлений водителю при бронировании экскурсии
# ======================
async def notify_drivers(order: dict):
    excursion = get_excursion_by_id(order["excursion_id"])
    start_time = excursion["start_time"] if excursion else "уточняется"

    route = json.loads(order["route"])
    route_text = " → ".join([p["address"] for p in route])

    text = (
        "🚐 <b>Новая экскурсия</b>\n\n"
        f"📍 <b>Экскурсия:</b> {order['excursion']}\n"
        f"📅 <b>Дата:</b> {order['date']}\n"
        f"⏰ <b>Время:</b> {start_time}\n\n"
        f"📌 <b>Место подачи:</b>\n"
        f"{order['pickup_address']}\n\n"
        f"🗺 <b>Маршрут:</b>\n"
        f"{route_text}\n\n"
        f"👥 <b>Пассажиры:</b> {order['count']}\n"
        f"📞 <b>Клиент:</b> {order['name']} | {order['phone']}\n"
        f"💰 <b>Сумма:</b> {order['price']} ₽"
    )

    for driver in DRIVERS:
        try:
            await bot.send_message(
                driver["telegram_id"],
                text,
                parse_mode="HTML"
            )
        except Exception as e:
            logging.error(
                f"Не удалось отправить уведомление водителю "
                f"{driver['name']} ({driver['telegram_id']}): {e}"
            )
# ======================
# Helper для сохранения заказов в эксель
# ======================

def save_order_to_excel(order: dict):
    file_exists = os.path.exists(EXCEL_FILE)

    if file_exists:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Дата бронирования",
            "ID бронирования",
            "Экскурсия",
            "Дата экскурсии",
            "Время",
            "ФИО",
            "Телефон",
            "Кол-во чел",
            "Сумма",
            "Предоплата",
            "Место подачи",
            "Маршрут",
            "Статус"
        ])

    ws.append([
        datetime.now().strftime("%d.%m.%Y %H:%M"),
        order["booking_id"],
        order["excursion"],
        order["date"],
        order.get("start_time", ""),
        order["name"],
        order["phone"],
        order["count"],
        order["price"],
        order.get("prepayment", 0),  # ← пока всегда 0
        order["pickup_address"],
        " → ".join([p["address"] for p in json.loads(order["route"])]),
        order["order_status"] # ← "Создан" или "Подписан"
    ])

    wb.save(EXCEL_FILE)

# ======================
# ХЕНДЛЕРЫ
# ======================

@dp.message(Command("start"))
async def start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(text=(
            f"Добро пожаловать 👋\n"
            f"Здесь вы можете выбрать и забронировать эскурсию по Москве на комфортном минивэне. Данный сервис работает ежедневно с 10:00 до 22:00\n"
        ), reply_markup=main_kb)

@dp.message(lambda m: m.text == "⬅️ Назад")
async def go_back(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Главное меню", reply_markup=main_kb)

@dp.message(lambda m: m.text == "🚗 Выбрать экскурсию")
async def choose_excursion(message: Message, state: FSMContext):
    await message.answer("Выберите экскурсию:", reply_markup=excursion_kb())
    await state.set_state(BookingStates.choose_excursion)

@dp.message(BookingStates.choose_excursion)
async def show_excursion(message: Message, state: FSMContext):
    selected = next((ex for ex in EXCURSIONS if ex["title"] == message.text), None)
    if not selected:
        await message.answer("Выберите экскурсию из списка.")
        return

    await state.update_data(
        excursion_id=selected["id"],
        excursion=selected["title"],
        start_time=selected["start_time"],  # 🔥 ВАЖНО
        price_per_person=selected["price"],
        prepayment_percent=selected["prepayment_percent"],
        pickup_address=(
            f"{selected['pickup']['title']}, "
            f"{selected['pickup']['address']} "
            f"({selected['pickup']['gis']})"
        ),
        route=json.dumps(selected["route"], ensure_ascii=False)
    )

    images = selected["images"][:10]
    media = []
    for img in images[:-1]:
        media.append(InputMediaPhoto(media=FSInputFile(img)))

    media.append(
        InputMediaPhoto(
            #media=FSInputFile(selected["images"][-1]),
            media=FSInputFile(images[-1]),  # ✅ ТОЛЬКО images
            caption=(
                f"<b>{selected['title']}</b>\n\n"
                f"{selected['description']}"
            ),
            parse_mode="HTML"
        )
    )

    await message.answer_media_group(media)
    await message.answer("Готовы забронировать?", reply_markup=book_kb())
    #await message.answer(selected["description"], reply_markup=book_kb())

@dp.callback_query(lambda c: c.data == "start_booking")
async def start_booking(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    now = date.today()

    # ⚠️ предполагается, что excursion_id уже сохранён в state ранее
    excursion_id = data.get("excursion_id")

    # 🔹 получаем доступные даты из БД (на 14 дней вперёд)
    dates = get_available_dates_range(
        excursion_id=excursion_id,
        start_date=now,
        days_ahead=14
    )

    await state.update_data(
        cal_year=now.year,
        cal_month=now.month
    )

    text = (
        "📅 <b>Выберите дату экскурсии</b>\n\n"
        "Бронирование доступно на 14 дней вперёд\n\n"
        f"Текущий месяц: <b>{month_title(now.year, now.month)}</b>\n\n"
        f"🟢 4-5 мест;\n 🟡 2–3 места;\n 🔴 1 место;\n ❌ нет мест"

    )

    await callback.message.answer(
        text,
        parse_mode="HTML",
        reply_markup=build_calendar(
            now.year,
            now.month,
            dates,                      # ← ВАЖНО: передаём dates
            blocked_dates=BLOCKED_DATES
        )
    )

    await state.set_state(BookingStates.date)
    await callback.answer()



# ======================
# 🔥 ВЫБОР ДАТЫ (ИЗМЕНЕНО)
# ======================

@dp.callback_query(lambda c: c.data.startswith("date:"))
async def select_date(callback: CallbackQuery, state: FSMContext):
    selected_date = callback.data.split(":")[1]
    today = date.today()
    max_date = today + timedelta(days=MAX_DAYS_AHEAD)

    chosen = datetime.strptime(selected_date, "%Y-%m-%d").date()

    if chosen < today or chosen > max_date:
        await callback.answer(
            "❌ Эту дату выбрать нельзя",
            show_alert=True
        )
        return

    await state.update_data(date=selected_date)
    await callback.message.answer("Введите ваше ФИО:")
    await state.set_state(BookingStates.name)
    await callback.answer()

# ======================
# 🔥 ПЕРЕКЛЮЧЕНИЕ МЕСЯЦЕВ (ИЗМЕНЕНО)
# ======================

@dp.callback_query(lambda c: c.data.startswith("cal_prev"))
async def calendar_prev(callback: CallbackQuery, state: FSMContext):
    _, year, month = callback.data.split(":")
    year, month = int(year), int(month)

    month -= 1
    if month == 0:
        month = 12
        year -= 1

    text = (
        "📅 <b>Выберите дату экскурсии</b>\n"
        f"Текущий месяц: <b>{month_title(year, month)}</b>\n\n"
        "🟢 много мест · 🟡 средняя загрузка · 🟠 1 место · ❌ нет мест"
    )

    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=build_calendar(
            year,
            month,
            blocked_dates=BLOCKED_DATES
        )
    )

    await callback.answer()


@dp.callback_query(lambda c: c.data.startswith("cal_next"))
async def calendar_next(callback: CallbackQuery, state: FSMContext):
    _, year, month = callback.data.split(":")
    year, month = int(year), int(month)

    month += 1
    if month == 13:
        month = 1
        year += 1

    text = (
        "📅 <b>Выберите дату экскурсии</b>\n"
        f"Текущий месяц: <b>{month_title(year, month)}</b>\n\n"
        "🟢 много мест · 🟡 средняя загрузка · 🟠 1 место · ❌ нет мест"
    )

    await callback.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=build_calendar(
            year,
            month,
            blocked_dates=BLOCKED_DATES
        )
    )

    await callback.answer()

#@dp.message(BookingStates.date)
#async def book_date(message: Message, state: FSMContext):
#    await state.update_data(date=message.text)
#    await message.answer("Введите ваше ФИО:")
#    await state.set_state(BookingStates.name)

@dp.message(BookingStates.name)
async def book_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("Введите номер телефона:")
    await state.set_state(BookingStates.phone)

@dp.message(BookingStates.phone)
async def book_phone(message: Message, state: FSMContext):
    await state.update_data(phone=message.text)
    await message.answer("Введите количество человек:")
    await state.set_state(BookingStates.count)

@dp.message(BookingStates.count)
async def book_count(message: Message, state: FSMContext):
    count = int(message.text)
    data = await state.get_data()

    booking_id = str(uuid.uuid4())
    total_price = data["price_per_person"] * count
    prepayment_amount = int(total_price * data["prepayment_percent"] / 100)

    success = book_places(
        data["excursion_id"],
        data["date"],
        count
    )

    if not success:
        await message.answer(
            "❌ Недостаточно свободных мест на выбранную дату.\n"
            "Пожалуйста, выберите другую дату."
        )
        return

    order_data = {
        "booking_id": booking_id,
        "tg_id": message.from_user.id,
        "name": data["name"],
        "phone": data["phone"],
        "pickup_address": data["pickup_address"],
        "excursion_id": data["excursion_id"],
        "excursion": data["excursion"],
        "date": data["date"],
        "start_time": data["start_time"],  # 🔥 НОВОЕ
        "count": count,
        "price": total_price,
        "route": data["route"],
        "order_status": "Создан",  # 🔹 ОБЯЗАТЕЛЬНО
        "prepayment": 0  # 🔹 поле для предоплаты
    }

    save_order(order_data) # Сохранение в БД
    save_order_to_excel(order_data)  # 🔥 Сохранение в Excel

    await message.answer(
        text=(
            f"Бронирование почти завершено\n"
            f"✅ Экскурсия: {data['excursion']}\n"
            f"✅ Дата: {data['date']}\n"
            f"✅ Кол-во человек: {count}\n"
            f"💵 Итоговая сумма к оплате: {total_price} ₽"
        ),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💰 Оплата на месте", callback_data="pay_on_place")]
        ])
    )

    await state.set_state(BookingStates.payment)

@dp.callback_query(lambda c: c.data == "paid")
async def paid(callback: CallbackQuery):
    booking = get_last_booking_by_user(callback.from_user.id)
    mark_paid(booking["booking_id"], booking["prepayment"])
    await callback.message.answer(
        "Оплата получена ✅\nТеперь вы можете ознакомиться с договором.",
        reply_markup=view_contract_kb
    )
    await callback.answer()

@dp.callback_query(lambda c: c.data == "pay_on_place")
async def pay_on_place(callback: CallbackQuery):
    booking = get_last_booking_by_user(callback.from_user.id)
    mark_paid(booking["booking_id"], 0)



    await callback.message.answer(
        "Вы выбрали оплату на месте ✅\n"
        #"Водитель уведомлён.\n\n"
        "Теперь вы можете ознакомиться с договором.",
        reply_markup=view_contract_kb
    )
    await callback.answer()

@dp.callback_query(lambda c: c.data == "view_contract")
async def view_contract_handler(callback: CallbackQuery):
    booking = get_last_booking_by_user(callback.from_user.id)
    pdf = generate_contract_pdf(booking, signed=False)
    await callback.message.answer_document(FSInputFile(pdf))
    await callback.message.answer("После ознакомления вы можете подписать договор:", reply_markup=sign_contract_kb)
    await callback.answer()

@dp.callback_query(lambda c: c.data == "sign_contract")
async def sign_contract_handler(callback: CallbackQuery):
    booking = get_last_booking_by_user(callback.from_user.id)
    sign_contract(booking["booking_id"])
    booking["order_status"] = "Подписан"
    booking["prepayment"] = 0

    save_order_to_excel(booking)
    pdf = generate_contract_pdf(booking, signed=True)
    await callback.message.answer_document(FSInputFile(pdf))

    excursion = get_excursion_by_id(booking["excursion_id"])
    start_time = excursion["start_time"] if excursion else "уточняется"

    await notify_drivers(booking)  # 🔥 Отправка уведомления водителю
    await callback.message.answer(
        text=(
            "✅ <b>Договор подписан. Водителю направлена информация по Вашему заказу</b>\n\n"
            f"📅 <b>Дата экскурсии:</b> {booking['date']}\n"
            f"⏰ <b>Время начала:</b> {start_time}\n\n"
            f"📍 <b>Место подачи:</b>\n"
            f"{booking['pickup_address']} https://go.2gis.com/2Q7no\n\n"
            "Пожалуйста, приезжайте за 10–15 минут до начала экскурсии.\n"
            "Будем рады видеть вас! 😊"
        ),
        parse_mode="HTML"
    )

    await callback.answer()

@dp.callback_query(lambda c: c.data == "ignore")
async def ignore_callback(callback: CallbackQuery):
    await callback.answer()

@dp.message(lambda m: m.text == "❓ FAQ")
async def faq(message: Message):
    await message.answer(FAQ_TEXT, parse_mode="HTML")

@dp.message(lambda m: m.text == "ℹ️ О нас")
async def about(message: Message):
    photo = FSInputFile("images/avatar.jpg")

    await message.answer_photo(
        photo=photo,
        caption=ABOUT_TEXT,
        parse_mode="HTML"
    )

@dp.message(lambda m: m.text == "⭐ Отзывы")
async def reviews(message: Message):
    for review in REVIEWS:
        photo = FSInputFile(review["photo"])
        await message.answer_photo(
            photo=photo,
            caption=review["text"],
            parse_mode="HTML"
        )

@dp.message(lambda m: m.text == "📞 Связаться")
async def contact(message: Message):
    await message.answer(CONTACT_TEXT, parse_mode="HTML")

# ======================
# ЗАПУСК
# ======================

async def main():
    init_db()
    now = datetime.now()
    for ex in EXCURSIONS:
        init_calendar_for_month(ex["id"], now.year, now.month)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())